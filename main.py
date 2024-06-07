import json
import os
import time
from pprint import pprint

import openpyxl
import requests
from loguru import logger

from api_wb import parser_wb_cards, push_media_in_comp, join_cards
from bd import Article, ArticleCharacteristicValue, Characteristic, CharacteristicValue
from config import headers


def create_card_wb(arts, subjectID):
    url_create_card = 'https://suppliers-api.wildberries.ru/content/v2/cards/upload'
    data_push = None
    for article in arts:
        try:
            data_push = [
                {
                    "subjectID": subjectID,
                    "variants": [
                        {
                            "vendorCode": article.art,
                            "title": article.name,
                            "description": article.descriptions,
                            "brand": article.brand,
                            "dimensions": {
                                "length": article.length,
                                "width": article.width,
                                "height": article.height,
                            },
                            "characteristics": get_article_characteristics(article.art),
                            "sizes": [
                                {
                                    "price": article.price,
                                    "skus": [
                                        f"{article.barcode}"
                                    ]
                                }
                            ]
                        }
                    ]
                }
            ]
        except Exception as ex:
            logger.error(ex)

        if data_push:
            data_json = json.dumps(data_push)
            try:
                response = requests.post(url_create_card, headers=headers, data=data_json)
                if response.status_code == 200:
                    logger.success(f'Создан артикул: {article.art}')
                elif response.status_code == 429:
                    logger.error(f'{response.text} {article.art}')
                    time.sleep(60)
                else:
                    logger.error(response.status_code)
                    logger.error(response.text)
            except Exception as ex:
                logger.error(ex)
            else:
                article.created_wb = True
                article.save()
                time.sleep(1)
        else:
            logger.error(f'Не удалось получить данные о артикуле {article.art}')


def get_article_characteristics(art):
    query = (ArticleCharacteristicValue
             .select(Article.art, Characteristic.id_wb, CharacteristicValue.value)
             .join(Article)
             .switch(ArticleCharacteristicValue)
             .join(CharacteristicValue)
             .join(Characteristic)
             .where(Article.art == art))

    characteristics = []
    for article_char_val in query:
        value = article_char_val.characteristic_value.value
        # Разбиваем значение по запятым, если они присутствуют
        if ',' in value:
            value = [v.strip() for v in value.split(',')]
        characteristic_data = {
            "id": article_char_val.characteristic_value.characteristic.id_wb,
            "value": value
        }
        characteristics.append(characteristic_data)

    return characteristics


def get_grouped_articles():
    query = (Article
             .select(Article.join_group, Article.art_wb, Article.imtID)
             .where(Article.joined == False)
             .order_by(Article.join_group))

    grouped_articles = {}
    for article in query:
        if article.join_group not in grouped_articles:
            grouped_articles[article.join_group] = {"arts": [], "imtIDs": []}
        grouped_articles[article.join_group]["arts"].append(article.art_wb)
        grouped_articles[article.join_group]["imtIDs"].append(article.imtID)

    return grouped_articles


def main():
    # Заводим карточки

    # articles = Article.select().where(Article.created_wb == False)
    # print(len(articles))
    # create_card_wb(articles, 1559)
    # time.sleep(1)

    # Получаем артикула вб и imtID
    # parser_wb_cards()
    # with open('dir_wb/wildberries_data_cards.json', 'r', encoding='utf-8') as f:
    #     data_arts_wb_in_file = json.load(f)
    # for item in data_arts_wb_in_file:
    #     try:
    #         art = Article.get(Article.art == item['vendorCode'])
    #     except Exception as ex:
    #         logger.error(ex)
    #     else:
    #         art.art_wb = item['nmID']
    #         art.imtID = item['imtID']
    #         art.save()
    # time.sleep(1)

    # Загружаем фото
    # articles = Article.select().where((Article.download_images == False) & (Article.created_wb == True))
    #
    # logger.info(f'Артикулов для загрузки фото: {len(articles)}')
    # for index_main, art in enumerate(articles, start=1):
    #     count = 0
    #     try:
    #         art_id = art.art_wb
    #         directory = art.folder_images_for_download
    #         image_list = []
    #         for file in os.listdir(directory):
    #             if 'Подложка' in file:
    #                 image_list.append(file)
    #
    #         for file in os.listdir(directory):
    #             if file not in image_list:
    #                 image_list.append(file)
    #         for index, file_name in enumerate(image_list, start=1):
    #             if push_media_in_comp(art_id, index, file_name, os.path.join(directory, file_name), art):
    #                 count += 1
    #                 time.sleep(0.5)
    #         else:
    #             art.download_images = True
    #             art.save()
    #         logger.debug(f'{index_main} Успешных загрузок для артикула {art}: {count}')
    #     except Exception as ex:
    #         logger.error(ex)

    # Склейка
    grouped_articles = get_grouped_articles()
    for key, value in grouped_articles.items():
        print(key, value['imtIDs'][0], value['arts'])
        try:
            join_cards(value['imtIDs'][0], value['arts'])
        except Exception as ex:
            logger.error(ex)
        else:
            arts = Article.select().where(Article.join_group == key)
            for art in arts:
                art.joined = True
                art.save()
        time.sleep(1)


def create_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Articles"

    # Добавление заголовков
    ws.append(["Склейка", "Артикул продавца", "Артикул WB", "Ссылка на задачу"])

    # Извлечение данных из базы данных и добавление их в лист
    query = Article.select(Article.join_group, Article.art, Article.art_wb, Article.url_task)
    for article in query:
        row = [article.join_group, article.art, article.art_wb]

        # Добавление гиперссылки
        cell_value = article.url_task
        ws.append(row + [cell_value])
        cell = ws.cell(row=ws.max_row, column=4)
        cell.hyperlink = article.url_task
        cell.style = "Hyperlink"
        cell.value = cell_value

    # Сохранение книги Excel
    wb.save("articles.xlsx")

if __name__ == '__main__':
    create_excel()
